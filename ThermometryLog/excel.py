"""

Работа с Excel.
Импорт и экспорт данных.

Обрабатываются листы вида:
|   Журнал термометрии в группе <group>    |
| ФИО    |     <date>     |     <date>     |
| <name> | <temperature>  | <temperature>  |
| <name> | <temperature>  | <temperature>  |
| <name> | <temperature>  | <temperature>  |
...

Название листа - название группы.

"""

from datetime import datetime
from typing import List, Tuple

import openpyxl
from openpyxl.styles import Font, Alignment

from database import ThermometryLog, Float
from logger import logger


def import_data(filename: str, database: ThermometryLog, group: int = 0):
    """
    Импортирует данные из книги Excel.
    :param filename: Имя файла.
    :param database: API для работы с базой данных.
    :param group: Группа, в которую импортируем данные.
    """

    logger.info("Загрузка книги.")
    workbook = openpyxl.load_workbook(filename)  # Открываем книгу
    sheet = workbook[workbook.sheetnames[0]]  # Выбираем лист

    logger.info("Импорт записей.")
    rows = sheet.max_row  # Кол-во строк в таблице
    cols = sheet.max_column  # Кол-во колонок в таблице

    if rows < 3 or cols < 2:
        return

    for i in range(3, rows + 1):
        # Валидация поля `name`
        if len(name := sheet.cell(i, 1).value) == 0:
            continue
        for j in range(2, cols + 1):
            try:  # Валидация поля `temperature`
                temperature = float(sheet.cell(i, j).value)
                if temperature == 0:
                    continue
            except ValueError:
                continue
            try:  # Валидация поля `date`
                date = datetime.strptime(sheet.cell(2, j).value, "%d.%m.%Y")
            except ValueError:
                continue

            database.insert(
                need_commit=False,
                name=name,
                temperature=Float(round(temperature, 1)),
                date=date.strftime("%d.%m.%Y"),
                grp=group,
                arrival_time=datetime.now().strftime("%H:%M"),
            )
    database.commit()


def export_data(
    filename: str,
    dates: List[str],
    database: ThermometryLog,
    group: Tuple[int, str] = (0, "Общая"),
):
    """
    Экспортирует данные в книгу Excel.
    :param filename: Имя файла.
    :param dates: Даты — названия страниц (в формате '%d.%m.%Y').
    :param database: API для работы с базой данных.
    :param group: Информация о группе. (<ID>, <название>)
    """

    logger.info("Создание книги.")
    workbook = openpyxl.Workbook()  # Создаем файл
    workbook.remove(workbook["Sheet"])  # Удаляем начальный лист

    logger.info("Заполнение книги.")
    logs: List[Tuple[int, str, float, str, int, str]] = database.sql.filter(
        database.table_name,
        date_egt=dates[0],
        date_elt=dates[-1],
        **({} if group[0] == 0 else {"grp": group[0]}),
    )
    data = {}  # {'<name>': {'<date>': <temperature>, ...}, ...}
    for obj in logs:
        if obj[1] not in data:
            data[obj[1]] = {}
        data[obj[1]][obj[3]] = obj[2]

    sheet = workbook.create_sheet(group[1])  # Новый лист
    # Заполняем ячейки
    sheet["A1"] = f"Журнал термометрии в группе `{group[1]}`"
    sheet["A1"].font = Font(bold=True)
    sheet["A1"].alignment = Alignment(horizontal="center")
    sheet.append(["ФИО", *dates])
    # Размер колонок
    sheet.column_dimensions["A"].width = 18
    columns = [sheet.cell(1, x).coordinate for x in range(2, len(dates) + 2)]
    for column in columns:
        sheet.column_dimensions[column[0]].width = 12
    # Объединяем ячейки
    sheet.merge_cells(f"A1:{columns[-1]}")

    for name in data:
        if len(name) == 0:
            continue
        temperatures = [
            (0 if date not in data[name] else data[name][date]) for date in dates
        ]
        sheet.append([name, *temperatures])

    workbook.save(filename)  # Сохраняем в файл
